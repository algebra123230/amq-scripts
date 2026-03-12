#!/usr/bin/env python3
"""
Download songs from an anime music Power Ranking Google Sheet as mp3.

Setup:
    ./install.sh

Usage:
    python download_pr.py <google_sheet_url> [--download-threads N] [--convert-threads M]
"""

import re
import sys
import platform
import subprocess
import io
import argparse
import threading
import queue
import time
from dataclasses import dataclass, field

from datetime import date
from pathlib import Path
from urllib.parse import urlparse

import requests
import openpyxl


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------

def parse_sheet_id(url: str) -> str:
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if match:
        return match.group(1)
    if re.match(r'^[a-zA-Z0-9_-]+$', url):
        return url
    raise ValueError(f"Could not parse Google Sheet ID from: {url}")


def download_xlsx(sheet_id: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print(f"Fetching sheet: {url}")
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return resp.content


def find_header_row(ws):
    for row in ws.iter_rows():
        values = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        if "id" in values and any("anime" in v for v in values):
            col_map = {str(cell.value).strip().lower(): cell.column
                       for cell in row if cell.value is not None}
            return row[0].row, col_map
    raise ValueError("Could not find header row (expected columns: ID, Anime Name, Song Info)")


def parse_song_info(text: str):
    """Parse '"Song Title" by Artist' into (title, artist)."""
    if not text:
        return "", ""
    m = re.match(r'^["\u201c\u300c](.+?)["\u201d\u300d]\s+by\s+(.+)$', text.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    parts = text.split(" by ", 1)
    if len(parts) == 2:
        return parts[0].strip().strip('"'), parts[1].strip()
    return text.strip(), ""


def get_cell_url(cell) -> str | None:
    if cell.hyperlink:
        target = cell.hyperlink.target
        if target:
            return target
    val = str(cell.value).strip() if cell.value else ""
    return val if val.startswith("http") else None


def find_mp3_column(col_map: dict) -> int | None:
    for key in col_map:
        if "mp3" in key or "audio" in key:
            return col_map[key]
    return None


def sanitize(name: str) -> str:
    name = re.sub(r'[/\\:*?"<>|]', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name.strip('_')


# ---------------------------------------------------------------------------
# Filesystem helpers
# ---------------------------------------------------------------------------

MIN_FILE_SIZE = 10 * 1024  # 10 KB


def already_downloaded(out_dir: Path, id_prefix: str) -> bool:
    """True if a reasonably-sized .mp3 with this ID prefix already exists."""
    return any(
        f.stat().st_size >= MIN_FILE_SIZE
        for f in out_dir.glob(f"{id_prefix}_*.mp3")
    )


def git_root() -> Path:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--show-toplevel"],
            capture_output=True, text=True, check=True
        )
        return Path(result.stdout.strip())
    except subprocess.CalledProcessError:
        return Path(__file__).parent.parent


# ---------------------------------------------------------------------------
# Download + convert (two separate stages for threading)
# ---------------------------------------------------------------------------

def download_raw(url: str, dest_base: Path) -> Path:
    """Download to dest_base.<ext> and return the path. No conversion."""
    is_youtube = "youtube.com" in url or "youtu.be" in url
    if is_youtube:
        output_template = str(dest_base) + ".%(ext)s"
        result = subprocess.run(
            ["yt-dlp", "-f", "bestaudio/best", "--no-playlist", "-q",
             "--remote-components", "ejs:github",
             "-o", output_template, url],
            capture_output=True, text=True, stdin=subprocess.DEVNULL,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or f"yt-dlp exited {result.returncode}")
        candidates = [
            f for f in dest_base.parent.iterdir()
            if f.stem == dest_base.name and f.suffix not in (".mp3", ".part")
        ]
        if not candidates:
            raise FileNotFoundError(f"yt-dlp produced no output file for {url}")
        return candidates[0]
    else:
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, stream=True, timeout=120, headers=headers)
        resp.raise_for_status()
        ext = Path(urlparse(url).path).suffix or ".webm"
        dest = Path(str(dest_base) + ext)
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)
        return dest


def convert_to_mp3(raw_file: Path, mp3_dest: Path):
    """Convert raw_file → mp3_dest with ffmpeg."""
    result = subprocess.run(
        ["ffmpeg", "-i", str(raw_file), "-vn", "-acodec", "libmp3lame", "-q:a", "2",
         "-y", str(mp3_dest)],
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or f"ffmpeg exited {result.returncode}")


# ---------------------------------------------------------------------------
# Worker threads
# ---------------------------------------------------------------------------

@dataclass
class Stats:
    secs: float = 0.0
    count: int = 0

_print_lock = threading.Lock()

def log(msg: str):
    with _print_lock:
        print(msg, flush=True)


_STOP = object()  # sentinel


def download_worker(dl_queue, conv_queue, dl_lock, dl_stats, errors, errors_lock):
    while True:
        item = dl_queue.get()
        if item is _STOP:
            dl_queue.task_done()
            break
        idx, n_total, base_name, media_url, mp3_dest = item
        log(f"[{idx}/{n_total}] ↓ {base_name}")
        dest_base = mp3_dest.with_suffix("")
        try:
            t0 = time.monotonic()
            raw_file = download_raw(media_url, dest_base)
            with dl_lock:
                dl_stats.secs += time.monotonic() - t0
                dl_stats.count += 1
            conv_queue.put((idx, n_total, base_name, raw_file, mp3_dest))
        except Exception as e:
            for f in dest_base.parent.iterdir():
                if f.stem == dest_base.name:
                    f.unlink(missing_ok=True)
            with errors_lock:
                errors.append((base_name, e))
        dl_queue.task_done()


def convert_worker(conv_queue, raw_files_lock, raw_files_converted, conv_lock, conv_stats, errors, errors_lock):
    while True:
        item = conv_queue.get()
        if item is _STOP:
            conv_queue.task_done()
            break
        idx, n_total, base_name, raw_file, mp3_dest = item
        log(f"[{idx}/{n_total}] ♪ {base_name}")
        try:
            t0 = time.monotonic()
            convert_to_mp3(raw_file, mp3_dest)
            with conv_lock:
                conv_stats.secs += time.monotonic() - t0
                conv_stats.count += 1
            with raw_files_lock:
                raw_files_converted.append(raw_file)
        except Exception as e:
            with errors_lock:
                errors.append((f"(convert) {base_name}", e))
        conv_queue.task_done()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def fmt(secs: float) -> str:
    m, s = divmod(int(secs), 60)
    return f"{m}m{s:02d}s" if m else f"{s}s"


def fmt_avg(total_secs: float, count: int) -> str:
    if count == 0:
        return "—"
    return f"{total_secs / count:.1f}s/file"


def main():
    parser = argparse.ArgumentParser(description="Download PR songs as mp3.")
    parser.add_argument("sheet_url", help="Google Sheet URL or sheet ID")
    parser.add_argument("--download-threads", type=int, default=2, metavar="N",
                        help="Parallel download threads (default: 2)")
    parser.add_argument("--convert-threads", type=int, default=2, metavar="M",
                        help="Parallel ffmpeg conversion threads (default: 2)")
    parser.add_argument("--output-dir", metavar="DIR",
                        help="Output directory name under media/ (default: prompt)")
    args = parser.parse_args()

    sheet_id = parse_sheet_id(args.sheet_url)

    today = date.today().strftime("%Y-%m-%d")
    if args.output_dir:
        dir_name = args.output_dir
    else:
        dir_name = input(f"Output directory name [{today}]: ").strip() or today
    out_dir = git_root() / "media" / dir_name
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Saving to: {out_dir}")

    xlsx_bytes = download_xlsx(sheet_id)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    header_row, col_map = find_header_row(ws)
    print(f"Found headers: {list(col_map.keys())}")

    id_col = col_map.get("id")
    anime_col = next((col_map[k] for k in col_map if "anime" in k), None)
    song_info_col = next((col_map[k] for k in col_map if "song info" in k or "song title" in k), None)
    mp3_col = find_mp3_column(col_map)

    if not all([id_col, anime_col, song_info_col]):
        print(f"ERROR: Missing required columns. Found: {list(col_map.keys())}")
        sys.exit(1)

    print("mp3 column found — will use mp3 links." if mp3_col
          else "No mp3 column — will use Song Info links.")

    # Parse all songs
    songs = []
    for row in ws.iter_rows(min_row=header_row + 1):
        id_cell = row[id_col - 1]
        if id_cell.value is None:
            continue
        try:
            song_id = int(id_cell.value)
        except (ValueError, TypeError):
            continue
        anime = str(row[anime_col - 1].value or "").strip()
        song_info_cell = row[song_info_col - 1]
        title, artist = parse_song_info(str(song_info_cell.value or ""))
        media_url = get_cell_url(row[mp3_col - 1]) if mp3_col else get_cell_url(song_info_cell)
        songs.append((song_id, anime, title, artist, media_url))

    # Filter out already-done and no-URL songs
    pending = []
    skipped = 0
    for song_id, anime, title, artist, media_url in songs:
        id_prefix = f"{song_id:02d}"
        base_name = sanitize(f"{id_prefix}_{anime}_{title}_by_{artist}")
        mp3_dest = out_dir / f"{base_name}.mp3"
        if already_downloaded(out_dir, id_prefix):
            skipped += 1
        elif not media_url:
            log(f"  (no URL) {base_name}")
            skipped += 1
        else:
            pending.append((base_name, media_url, mp3_dest))

    n_total = len(pending)
    print(f"{len(songs)} songs found. {skipped} skipped, {n_total} to download.")
    if n_total == 0:
        print("Nothing to do.")
        return

    dl_queue: queue.Queue = queue.Queue()
    conv_queue: queue.Queue = queue.Queue()
    raw_files_converted: list[Path] = []
    raw_files_lock = threading.Lock()
    dl_lock = threading.Lock()
    dl_stats = Stats()
    conv_lock = threading.Lock()
    conv_stats = Stats()
    errors: list[tuple[str, Exception]] = []
    errors_lock = threading.Lock()

    # Enqueue pending songs
    for idx, (base_name, media_url, mp3_dest) in enumerate(pending, 1):
        dl_queue.put((idx, n_total, base_name, media_url, mp3_dest))

    t_start = time.monotonic()

    # Start convert workers first (they wait on conv_queue)
    conv_threads = [
        threading.Thread(target=convert_worker,
                         args=(conv_queue, raw_files_lock, raw_files_converted,
                               conv_lock, conv_stats, errors, errors_lock),
                         daemon=True)
        for _ in range(args.convert_threads)
    ]
    for t in conv_threads:
        t.start()

    # Start download workers
    dl_threads = [
        threading.Thread(target=download_worker,
                         args=(dl_queue, conv_queue, dl_lock, dl_stats,
                               errors, errors_lock),
                         daemon=True)
        for _ in range(args.download_threads)
    ]
    for t in dl_threads:
        t.start()

    # Signal download workers to stop after queue is drained
    for _ in dl_threads:
        dl_queue.put(_STOP)

    dl_queue.join()  # wait for all downloads (and their conv_queue.put calls) to finish

    # Signal convert workers to stop
    for _ in conv_threads:
        conv_queue.put(_STOP)

    conv_queue.join()  # wait for all conversions to finish

    t_end = time.monotonic()

    if raw_files_converted:
        total_mb = sum(f.stat().st_size for f in raw_files_converted if f.exists()) / (1024 * 1024)
        # ffmpeg modifies terminal settings even with redirected I/O; restore them before prompting
        if platform.system() != "Windows":
            subprocess.run(["stty", "sane"], check=False)
        prompt = f"\nDelete {len(raw_files_converted)} raw video files ({total_mb:.1f} MB)? [y/N] "
        answer = input(prompt).strip().lower()
        if answer == "y":
            for f in raw_files_converted:
                f.unlink(missing_ok=True)
            print("Raw files deleted.")
        else:
            print("Raw files kept.")

    nd, nc = dl_stats.count, conv_stats.count
    print(
        f"\nDownloaded {nd} file{'s' if nd != 1 else ''} in {fmt(dl_stats.secs)}"
        f" ({fmt_avg(dl_stats.secs, nd)})"
    )
    print(
        f"Converted  {nc} file{'s' if nc != 1 else ''} in {fmt(conv_stats.secs)}"
        f" ({fmt_avg(conv_stats.secs, nc)})"
    )
    print(f"End-to-end: {fmt(t_end - t_start)}")

    if errors:
        print(f"\n{len(errors)} download error{'s' if len(errors) != 1 else ''}:")
        for name, exc in errors:
            print(f"  {name}: {exc}")
    print("Done.")


if __name__ == "__main__":
    main()
